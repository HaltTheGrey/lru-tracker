"""Template management for bulk station import."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List, Tuple
from models import Station
from config import Colors
from validators import validate_station_name, validate_number
from logger import get_logger

logger = get_logger()


class TemplateManager:
    """Handles template creation and import."""
    
    def create_template(self, filename: str) -> None:
        """Generate Excel template for bulk station import."""
        wb = openpyxl.Workbook()
        ws: Worksheet = wb.active
        ws.title = "Station Setup"
        
        # Title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = "LRU Station Setup Template"
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color=Colors.HEADER_BG, end_color=Colors.HEADER_BG, fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Instructions
        ws.merge_cells('A2:E2')
        instructions = ws['A2']
        instructions.value = "Fill in your stations below. Do not modify the header row (row 3)."
        instructions.font = Font(italic=True)
        instructions.alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 25
        
        # Headers
        headers = ["Station Name", "Min LRU", "Max LRU", "Current LRU (Optional)", "Notes (Optional)"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(3, col_num)
            cell.value = header
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color=Colors.HEADER_SECONDARY, end_color=Colors.HEADER_SECONDARY, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[3].height = 25
        
        # Example data
        examples = [
            ["Pack Station 1", 5, 20, 10, "Main packing area"],
            ["Dock Door A", 10, 30, 15, "Receiving dock"],
            ["Induct Station 1", 3, 15, 8, "Induction line"],
        ]
        
        example_fill = PatternFill(start_color=Colors.EXAMPLE_BG, end_color=Colors.EXAMPLE_BG, fill_type="solid")
        for row_num, example in enumerate(examples, 4):
            for col_num, value in enumerate(example, 1):
                cell = ws.cell(row_num, col_num)
                cell.value = value
                cell.fill = example_fill
                if col_num in [2, 3, 4]:
                    cell.alignment = Alignment(horizontal='center')
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
        
        # Borders
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for row in range(3, 30):
            for col in range(1, 6):
                ws.cell(row, col).border = thin_border
        
        # Instructions sheet
        ws_inst = wb.create_sheet("Instructions")
        instructions_text = [
            ["LRU Station Template - Instructions"],
            [""],
            ["1. Fill in Station Information"],
            ["   • Station Name: Unique name for each station"],
            ["   • Min LRU: Minimum threshold (alerts when below)"],
            ["   • Max LRU: Maximum threshold (alerts when at/over)"],
            ["   • Current LRU: Optional starting count (defaults to 0)"],
            ["   • Notes: Optional description"],
            [""],
            ["2. Guidelines"],
            ["   • Do NOT modify the header row"],
            ["   • Station names must be unique"],
            ["   • Min must be less than Max"],
            ["   • Delete example rows before importing"],
            [""],
            ["3. Import Process"],
            ["   • Save this file after filling in data"],
            ["   • In LRU Tracker, click 'Import from Template'"],
            ["   • Select this file and confirm"],
        ]
        
        for row_num, (text,) in enumerate(instructions_text, 1):
            ws_inst.cell(row_num, 1).value = text
            if row_num == 1:
                ws_inst.cell(row_num, 1).font = Font(size=14, bold=True)
        
        ws_inst.column_dimensions['A'].width = 50
        
        wb.save(filename)
        logger.info(f"Template created: {filename}")
    
    def import_from_template(self, filename: str, existing_stations: Dict[str, Station]) -> Tuple[List[Station], List[str]]:
        """Import stations from template file. Returns (stations, errors)."""
        wb = openpyxl.load_workbook(filename)
        ws: Worksheet = wb["Station Setup"] if "Station Setup" in wb.sheetnames else wb.active
        
        imported_stations = []
        errors = []
        
        # Find header row
        header_row = None
        for row in range(1, 10):
            cell_value = ws.cell(row, 1).value
            if cell_value and "Station Name" in str(cell_value):
                header_row = row
                break
        
        if not header_row:
            raise ValueError("Could not find header row with 'Station Name'")
        
        # Read data
        for row_num in range(header_row + 1, ws.max_row + 1):
            station_name = ws.cell(row_num, 1).value
            min_lru = ws.cell(row_num, 2).value
            max_lru = ws.cell(row_num, 3).value
            current_lru = ws.cell(row_num, 4).value
            
            if not station_name or str(station_name).strip() == "":
                continue
            
            station_name = str(station_name).strip()
            
            # Validate
            if not validate_station_name(station_name):
                errors.append(f"Row {row_num}: Invalid station name '{station_name}'")
                continue
            
            min_valid, min_val = validate_number(str(min_lru) if min_lru else "5")
            max_valid, max_val = validate_number(str(max_lru) if max_lru else "20")
            current_valid, current_val = validate_number(str(current_lru) if current_lru else "0")
            
            if not min_valid or not max_valid:
                errors.append(f"Row {row_num}: Invalid min/max values for '{station_name}'")
                continue
            
            if min_val > max_val:
                errors.append(f"Row {row_num}: Min > Max for '{station_name}'")
                continue
            
            if station_name in existing_stations:
                errors.append(f"Row {row_num}: Station '{station_name}' already exists")
                continue
            
            station = Station(
                name=station_name,
                current=current_val if current_valid else 0,
                min_lru=min_val,
                max_lru=max_val
            )
            imported_stations.append(station)
        
        logger.info(f"Imported {len(imported_stations)} stations from template, {len(errors)} errors")
        return imported_stations, errors
