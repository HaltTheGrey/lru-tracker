"""Export functionality for Excel and CSV reports."""
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List
from datetime import datetime
from models import Station, GlobalHistoryEntry
from config import Colors, TIMESTAMP_FORMAT, FILE_TIMESTAMP_FORMAT


class ExcelColors:
    """Enhanced color scheme for Excel exports."""
    HEADER_PRIMARY = '1F4788'      # Deep blue
    HEADER_SECONDARY = '2E5C8A'    # Medium blue
    CRITICAL = 'E74C3C'            # Red
    WARNING = 'F39C12'             # Orange
    SUCCESS = '27AE60'             # Green
    INFO = '3498DB'                # Light blue
    NEUTRAL = 'ECF0F1'             # Light gray
    TEXT_DARK = '2C3E50'           # Dark gray
    BORDER_COLOR = 'BDC3C7'        # Gray border


class ExportManager:
    """Handles all export operations."""
    
    @staticmethod
    def create_header_style() -> tuple:
        """Create enhanced header styling."""
        fill = PatternFill(start_color=ExcelColors.HEADER_PRIMARY, end_color=ExcelColors.HEADER_PRIMARY, fill_type="solid")
        font = Font(color="FFFFFF", bold=True, size=13, name='Calibri')
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin', color=ExcelColors.BORDER_COLOR),
            right=Side(style='thin', color=ExcelColors.BORDER_COLOR),
            top=Side(style='thin', color=ExcelColors.BORDER_COLOR),
            bottom=Side(style='medium', color='FFFFFF')
        )
        return fill, font, alignment, border
    
    @staticmethod
    def get_status_fill(current: int, min_val: int, max_val: int) -> PatternFill:
        """Get enhanced status fill color based on values."""
        if current < min_val:
            return PatternFill(start_color=ExcelColors.CRITICAL, end_color=ExcelColors.CRITICAL, fill_type="solid")
        elif current >= max_val:
            return PatternFill(start_color=ExcelColors.WARNING, end_color=ExcelColors.WARNING, fill_type="solid")
        return PatternFill(start_color=ExcelColors.SUCCESS, end_color=ExcelColors.SUCCESS, fill_type="solid")
    
    @staticmethod
    def apply_cell_border(cell) -> None:
        """Apply consistent border to cell."""
        cell.border = Border(
            left=Side(style='thin', color=ExcelColors.BORDER_COLOR),
            right=Side(style='thin', color=ExcelColors.BORDER_COLOR),
            top=Side(style='thin', color=ExcelColors.BORDER_COLOR),
            bottom=Side(style='thin', color=ExcelColors.BORDER_COLOR)
        )
    
    def export_new_report(self, filename: str, stations: Dict[str, Station], 
                         history: List[GlobalHistoryEntry]) -> None:
        """Export enhanced Excel report with current status and history."""
        wb = openpyxl.Workbook()
        ws: Worksheet = wb.active
        ws.title = "Current Status"
        
        # Add title row
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"LRU Tracker Report - {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        title_cell.font = Font(size=16, bold=True, color=ExcelColors.HEADER_PRIMARY, name='Calibri')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Header
        headers = ["Station Name", "Current LRU", "Min", "Max", "Status", "Last Updated"]
        ws.append(headers)
        
        # Style header
        header_fill, header_font, header_align, header_border = self.create_header_style()
        for col in range(1, len(headers) + 1):
            cell = ws.cell(2, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = header_border
        ws.row_dimensions[2].height = 25
        
        # Data
        for name in sorted(stations.keys()):
            station = stations[name]
            status = station.get_status()
            last_updated = station.history[-1].timestamp if station.history else "Never"
            
            row = [name, station.current, station.min_lru, station.max_lru, status, last_updated]
            ws.append(row)
            row_num = ws.max_row
            
            # Style data cells
            for col in range(1, 7):
                cell = ws.cell(row_num, col)
                self.apply_cell_border(cell)
                cell.alignment = Alignment(horizontal='center' if col > 1 else 'left', vertical='center')
                cell.font = Font(name='Calibri', size=11)
                
                # Alternate row colors
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color=ExcelColors.NEUTRAL, end_color=ExcelColors.NEUTRAL, fill_type="solid")
            
            # Color code status with white text
            status_cell = ws.cell(row_num, 5)
            status_cell.fill = self.get_status_fill(station.current, station.min_lru, station.max_lru)
            status_cell.font = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 22
        
        # Freeze panes
        ws.freeze_panes = 'A3'
        
        # Create history sheet if there's history
        if history:
            self._add_history_sheet(wb, history)
        
        wb.save(filename)
    
    def _add_history_sheet(self, wb: openpyxl.Workbook, 
                          history: List[GlobalHistoryEntry]) -> None:
        """Add enhanced history sheet to workbook."""
        ws_history = wb.create_sheet("History")
        
        # Add title
        ws_history.merge_cells('A1:E1')
        title_cell = ws_history['A1']
        title_cell.value = "LRU Update History"
        title_cell.font = Font(size=16, bold=True, color=ExcelColors.HEADER_PRIMARY, name='Calibri')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_history.row_dimensions[1].height = 30
        
        history_headers = ["Station", "Timestamp", "Count", "Min", "Max"]
        ws_history.append(history_headers)
        
        # Style header
        header_fill, header_font, header_align, header_border = self.create_header_style()
        for col in range(1, len(history_headers) + 1):
            cell = ws_history.cell(2, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = header_border
        ws_history.row_dimensions[2].height = 25
        
        # Add history data
        for record in history:
            ws_history.append([
                record.station, record.timestamp, record.count,
                record.min_lru, record.max_lru
            ])
            row_num = ws_history.max_row
            
            # Style cells
            for col in range(1, 6):
                cell = ws_history.cell(row_num, col)
                self.apply_cell_border(cell)
                cell.alignment = Alignment(horizontal='center' if col > 1 else 'left', vertical='center')
                cell.font = Font(name='Calibri', size=11)
                
                # Alternate row colors
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color=ExcelColors.NEUTRAL, end_color=ExcelColors.NEUTRAL, fill_type="solid")
        
        # Adjust column widths
        ws_history.column_dimensions['A'].width = 30
        ws_history.column_dimensions['B'].width = 22
        ws_history.column_dimensions['C'].width = 15
        ws_history.column_dimensions['D'].width = 12
        ws_history.column_dimensions['E'].width = 12
        
        # Freeze panes
        ws_history.freeze_panes = 'A3'
    
    def append_to_existing(self, filename: str, stations: Dict[str, Station]) -> str:
        """Append enhanced snapshot to existing Excel file."""
        wb = openpyxl.load_workbook(filename)
        
        # Create new sheet with timestamp
        sheet_name = f"Snapshot_{datetime.now().strftime(FILE_TIMESTAMP_FORMAT)}"
        ws = wb.create_sheet(sheet_name)
        
        # Add title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"Snapshot - {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        title_cell.font = Font(size=16, bold=True, color=ExcelColors.HEADER_PRIMARY, name='Calibri')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Header
        headers = ["Station Name", "Current LRU", "Min", "Max", "Status", "Timestamp"]
        ws.append(headers)
        
        # Style header
        header_fill, header_font, header_align, header_border = self.create_header_style()
        for col in range(1, len(headers) + 1):
            cell = ws.cell(2, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = header_border
        ws.row_dimensions[2].height = 25
        
        # Data
        timestamp = datetime.now().strftime(TIMESTAMP_FORMAT)
        for name in sorted(stations.keys()):
            station = stations[name]
            status = station.get_status()
            
            row = [name, station.current, station.min_lru, station.max_lru, status, timestamp]
            ws.append(row)
            row_num = ws.max_row
            
            # Style cells
            for col in range(1, 7):
                cell = ws.cell(row_num, col)
                self.apply_cell_border(cell)
                cell.alignment = Alignment(horizontal='center' if col > 1 else 'left', vertical='center')
                cell.font = Font(name='Calibri', size=11)
                
                # Alternate row colors
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color=ExcelColors.NEUTRAL, end_color=ExcelColors.NEUTRAL, fill_type="solid")
            
            # Color code status
            status_cell = ws.cell(row_num, 5)
            status_cell.fill = self.get_status_fill(station.current, station.min_lru, station.max_lru)
            status_cell.font = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 22
        
        # Freeze panes
        ws.freeze_panes = 'A3'
        
        wb.save(filename)
        return sheet_name
    
    def create_trend_report(self, filename: str, station: Station) -> None:
        """Create professional trend report with enhanced chart and analysis for a station."""
        wb = openpyxl.Workbook()
        ws: Worksheet = wb.active
        
        # Shorten title if too long (Excel sheet name limit is 31 chars)
        sheet_title = station.name[:28] + "..." if len(station.name) > 31 else station.name
        ws.title = sheet_title
        
        # ===== TITLE SECTION =====
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"📈 LRU Trend Analysis: {station.name}"
        title_cell.font = Font(size=16, bold=True, color=ExcelColors.HEADER_PRIMARY, name='Calibri')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color=ExcelColors.NEUTRAL, end_color=ExcelColors.NEUTRAL, fill_type="solid")
        ws.row_dimensions[1].height = 35
        
        # ===== SUMMARY SECTION =====
        ws.merge_cells('A2:F2')
        summary_cell = ws['A2']
        summary_cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        summary_cell.font = Font(size=11, italic=True, color=ExcelColors.TEXT_DARK, name='Calibri')
        summary_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 20
        
        # ===== STATISTICS CARDS (Row 3) =====
        current_row = 3
        ws.row_dimensions[current_row].height = 25
        
        # Calculate statistics
        if station.history:
            counts = [record.count for record in station.history]
            avg_count = sum(counts) / len(counts)
            min_count = min(counts)
            max_count = max(counts)
            total_records = len(counts)
            current_status = station.get_status()
        else:
            avg_count = min_count = max_count = total_records = 0
            current_status = "No Data"
        
        # Create statistics cards
        stats_headers = ["Current LRU", "Min Threshold", "Max Threshold", "Average", "Status", "Data Points"]
        stats_values = [
            station.current,
            station.min_lru,
            station.max_lru,
            f"{avg_count:.1f}" if station.history else "N/A",
            current_status,
            total_records
        ]
        
        # Style statistics section
        for col_idx, (header, value) in enumerate(zip(stats_headers, stats_values), 1):
            # Header cell
            header_cell = ws.cell(current_row, col_idx)
            header_cell.value = header
            header_cell.font = Font(size=10, bold=True, color='FFFFFF', name='Calibri')
            header_cell.fill = PatternFill(start_color=ExcelColors.HEADER_SECONDARY, 
                                          end_color=ExcelColors.HEADER_SECONDARY, fill_type="solid")
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.border = Border(
                left=Side(style='thin', color=ExcelColors.BORDER_COLOR),
                right=Side(style='thin', color=ExcelColors.BORDER_COLOR),
                top=Side(style='thin', color=ExcelColors.BORDER_COLOR),
                bottom=Side(style='thin', color=ExcelColors.BORDER_COLOR)
            )
            
            # Value cell
            value_cell = ws.cell(current_row + 1, col_idx)
            value_cell.value = value
            value_cell.font = Font(size=12, bold=True, name='Calibri')
            value_cell.alignment = Alignment(horizontal='center', vertical='center')
            value_cell.border = Border(
                left=Side(style='thin', color=ExcelColors.BORDER_COLOR),
                right=Side(style='thin', color=ExcelColors.BORDER_COLOR),
                top=Side(style='thin', color=ExcelColors.BORDER_COLOR),
                bottom=Side(style='medium', color=ExcelColors.BORDER_COLOR)
            )
            
            # Color code based on status for Current LRU
            if col_idx == 1:  # Current LRU
                value_cell.fill = self.get_status_fill(station.current, station.min_lru, station.max_lru)
                value_cell.font = Font(size=12, bold=True, color='FFFFFF', name='Calibri')
            elif col_idx == 5:  # Status
                if current_status == "Critical":
                    value_cell.fill = PatternFill(start_color=ExcelColors.CRITICAL, 
                                                 end_color=ExcelColors.CRITICAL, fill_type="solid")
                    value_cell.font = Font(size=12, bold=True, color='FFFFFF', name='Calibri')
                elif current_status == "Warning":
                    value_cell.fill = PatternFill(start_color=ExcelColors.WARNING, 
                                                 end_color=ExcelColors.WARNING, fill_type="solid")
                    value_cell.font = Font(size=12, bold=True, color='FFFFFF', name='Calibri')
                elif current_status == "Good":
                    value_cell.fill = PatternFill(start_color=ExcelColors.SUCCESS, 
                                                 end_color=ExcelColors.SUCCESS, fill_type="solid")
                    value_cell.font = Font(size=12, bold=True, color='FFFFFF', name='Calibri')
            else:
                value_cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type="solid")
        
        ws.row_dimensions[current_row + 1].height = 30
        
        # ===== DATA TABLE SECTION =====
        current_row = 6
        
        # Add spacing row
        current_row += 1
        
        # Data headers
        headers = ["#", "Timestamp", "LRU Count", "Min", "Max", "Status", "Variance"]
        ws.append(headers)
        current_row += 1
        
        # Style data headers
        header_fill, header_font, header_align, header_border = self.create_header_style()
        for col in range(1, len(headers) + 1):
            cell = ws.cell(current_row, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = header_border
        ws.row_dimensions[current_row].height = 25
        
        # Add data rows with enhanced formatting
        for idx, record in enumerate(station.history, 1):
            variance = record.count - avg_count if station.history else 0
            status = "Critical" if record.count < station.min_lru else \
                    "Warning" if record.count >= station.max_lru else "Good"
            
            row_data = [idx, record.timestamp, record.count, station.min_lru, 
                       station.max_lru, status, f"{variance:+.1f}"]
            ws.append(row_data)
            row_num = ws.max_row
            
            # Style data cells
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row_num, col)
                self.apply_cell_border(cell)
                cell.alignment = Alignment(horizontal='center' if col > 1 else 'right', vertical='center')
                cell.font = Font(name='Calibri', size=10)
                
                # Alternate row colors
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type="solid")
            
            # Color code status column
            status_cell = ws.cell(row_num, 6)
            if status == "Critical":
                status_cell.fill = PatternFill(start_color=ExcelColors.CRITICAL, 
                                              end_color=ExcelColors.CRITICAL, fill_type="solid")
                status_cell.font = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
            elif status == "Warning":
                status_cell.fill = PatternFill(start_color=ExcelColors.WARNING, 
                                              end_color=ExcelColors.WARNING, fill_type="solid")
                status_cell.font = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
            else:
                status_cell.fill = PatternFill(start_color=ExcelColors.SUCCESS, 
                                              end_color=ExcelColors.SUCCESS, fill_type="solid")
                status_cell.font = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
            
            # Color code variance (positive = green, negative = red)
            variance_cell = ws.cell(row_num, 7)
            if variance > 0:
                variance_cell.font = Font(bold=True, color=ExcelColors.SUCCESS, size=10, name='Calibri')
            elif variance < 0:
                variance_cell.font = Font(bold=True, color=ExcelColors.CRITICAL, size=10, name='Calibri')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8   # #
        ws.column_dimensions['B'].width = 22  # Timestamp
        ws.column_dimensions['C'].width = 12  # LRU Count
        ws.column_dimensions['D'].width = 10  # Min
        ws.column_dimensions['E'].width = 10  # Max
        ws.column_dimensions['F'].width = 12  # Status
        ws.column_dimensions['G'].width = 12  # Variance
        
        # Freeze panes (freeze title, summary, and header rows)
        ws.freeze_panes = 'A9'
        
        # ===== CREATE PROFESSIONAL CHART =====
        if station.history and len(station.history) > 1:
            chart = LineChart()
            chart.title = f"LRU Trend Analysis"
            chart.style = 12  # Professional style
            chart.y_axis.title = "LRU Count"
            chart.x_axis.title = "Timeline"
            chart.height = 12  # Taller chart
            chart.width = 20   # Wider chart
            
            # Add LRU Count line (bold)
            data_ref = Reference(ws, min_col=3, max_col=3, min_row=8, max_row=8 + len(station.history))
            chart.add_data(data_ref, titles_from_data=True)
            
            # Add Min threshold line
            min_ref = Reference(ws, min_col=4, max_col=4, min_row=8, max_row=8 + len(station.history))
            chart.add_data(min_ref, titles_from_data=True)
            
            # Add Max threshold line
            max_ref = Reference(ws, min_col=5, max_col=5, min_row=8, max_row=8 + len(station.history))
            chart.add_data(max_ref, titles_from_data=True)
            
            # Set categories (timestamps)
            cats = Reference(ws, min_col=2, min_row=9, max_row=8 + len(station.history))
            chart.set_categories(cats)
            
            # Customize line colors
            chart.series[0].graphicalProperties.line.solidFill = ExcelColors.INFO  # LRU Count - Blue
            chart.series[0].graphicalProperties.line.width = 3  # Thicker line
            chart.series[1].graphicalProperties.line.solidFill = ExcelColors.CRITICAL  # Min - Red
            chart.series[1].graphicalProperties.line.width = 2
            chart.series[1].graphicalProperties.line.dashStyle = "dash"  # Dashed line
            chart.series[2].graphicalProperties.line.solidFill = ExcelColors.WARNING  # Max - Orange
            chart.series[2].graphicalProperties.line.width = 2
            chart.series[2].graphicalProperties.line.dashStyle = "dash"  # Dashed line
            
            # Place chart to the right of the data
            ws.add_chart(chart, "I3")
        
        wb.save(filename)
