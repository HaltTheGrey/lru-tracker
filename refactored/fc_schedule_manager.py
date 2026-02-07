"""FC Schedule import/export functionality."""
import csv
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Tuple
from collections import defaultdict
from datetime import datetime
from models import Station
from config import ALL_TIME_SLOTS, TIME_SLOT_MAP, TIMESTAMP_FORMAT
from logger import get_logger

logger = get_logger()


class ExcelColors:
    """Color scheme for FC Schedule exports."""
    HEADER_BLUE = '1F4788'
    SHIFT1_HEADER = '2E5C8A'
    SHIFT2_HEADER = '34495E'
    CRITICAL = 'E74C3C'
    WARNING = 'F39C12'
    SUCCESS = '27AE60'
    INFO = '3498DB'
    NEUTRAL = 'ECF0F1'
    WHITE = 'FFFFFF'
    TEXT_DARK = '2C3E50'
    BORDER_COLOR = 'BDC3C7'
    LIGHT_BLUE = 'D6EAF8'
    LIGHT_GREEN = 'D5F4E6'


class FCScheduleManager:
    """Handles FC Standard Work Spreadsheet format."""
    
    def import_from_csv(self, filename: str, existing_stations: Dict[str, Station]) -> Tuple[List[Station], List[str]]:
        """Import stations from FC schedule CSV. Returns (stations, errors)."""
        imported_stations = []
        errors = []
        
        with open(filename, 'r', encoding='utf-8-sig') as f:
            csv_reader = csv.reader(f)
            rows = list(csv_reader)
        
        for row_num, row in enumerate(rows, 1):
            if len(row) < 3:
                continue
            
            lru_name = str(row[0]).strip() if row[0] else ""
            test_desc = str(row[1]).strip() if row[1] else ""
            rack_location = str(row[2]).strip() if row[2] else ""
            
            # Skip headers
            if not lru_name or "LRU" in lru_name or "Shift" in lru_name:
                continue
            
            # Create station name
            station_name = f"{lru_name} - {rack_location}" if rack_location else lru_name
            
            # Extract batch size from test description (B=X format)
            batch_match = re.search(r'B\s*=\s*(\d+)', test_desc, re.IGNORECASE)
            
            if batch_match:
                batch_size = int(batch_match.group(1))
                min_val = max(1, batch_size // 2)
                max_val = batch_size * 2
            else:
                min_val = 5
                max_val = 20
            
            if station_name in existing_stations:
                errors.append(f"Row {row_num}: Station '{station_name}' already exists")
                continue
            
            station = Station(
                name=station_name,
                current=0,
                min_lru=min_val,
                max_lru=max_val,
                test_description=test_desc,
                rack_location=rack_location
            )
            imported_stations.append(station)
        
        logger.info(f"Imported {len(imported_stations)} stations from FC schedule, {len(errors)} errors")
        return imported_stations, errors
    
    def export_to_csv(self, filename: str, stations: Dict[str, Station]) -> None:
        """Export stations in professional FC schedule Excel format with enhanced styling."""
        # Check if we should export as Excel (recommended) or CSV
        is_excel = filename.lower().endswith('.xlsx') or filename.lower().endswith('.xls')
        
        if is_excel:
            self._export_to_excel(filename, stations)
        else:
            self._export_to_csv_legacy(filename, stations)
    
    def _export_to_excel(self, filename: str, stations: Dict[str, Station]) -> None:
        """Export stations in professional Excel format with styling and formulas."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FC Schedule"
        
        # ===== TITLE SECTION =====
        ws.merge_cells('A1:O1')
        title_cell = ws['A1']
        title_cell.value = f"🏭 FC Standard Work Schedule - {datetime.now().strftime('%B %d, %Y')}"
        title_cell.font = Font(size=16, bold=True, color=ExcelColors.HEADER_BLUE, name='Calibri')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color=ExcelColors.NEUTRAL, end_color=ExcelColors.NEUTRAL, fill_type="solid")
        ws.row_dimensions[1].height = 35
        
        # ===== INSTRUCTIONS ROW =====
        ws.merge_cells('A2:O2')
        instructions = ws['A2']
        instructions.value = "Record the number of batches scheduled for each time slot. Update counts throughout the shift."
        instructions.font = Font(size=10, italic=True, color=ExcelColors.TEXT_DARK, name='Calibri')
        instructions.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 20
        
        current_row = 3
        
        # ===== SHIFT HEADERS (Row 3) =====
        # Merge cells for shift labels
        ws.merge_cells(f'D{current_row}:I{current_row}')
        shift1_cell = ws[f'D{current_row}']
        shift1_cell.value = "1st Shift - Record # of Batches to Schedule"
        shift1_cell.font = Font(size=12, bold=True, color=ExcelColors.WHITE, name='Calibri')
        shift1_cell.fill = PatternFill(start_color=ExcelColors.SHIFT1_HEADER, 
                                       end_color=ExcelColors.SHIFT1_HEADER, fill_type="solid")
        shift1_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells(f'J{current_row}:L{current_row}')
        shift2_cell = ws[f'J{current_row}']
        shift2_cell.value = "2nd Shift - Record # of Batches to Schedule"
        shift2_cell.font = Font(size=12, bold=True, color=ExcelColors.WHITE, name='Calibri')
        shift2_cell.fill = PatternFill(start_color=ExcelColors.SHIFT2_HEADER, 
                                       end_color=ExcelColors.SHIFT2_HEADER, fill_type="solid")
        shift2_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        # ===== COLUMN HEADERS (Row 4) =====
        headers = ['LRU', 'Test to Schedule', 'Rack Location'] + ALL_TIME_SLOTS + ['Total', 'Status']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(current_row, col_idx)
            cell.value = header
            cell.font = Font(size=11, bold=True, color=ExcelColors.WHITE, name='Calibri')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Color code by shift
            if col_idx <= 3:  # LRU, Test, Rack Location
                cell.fill = PatternFill(start_color=ExcelColors.HEADER_BLUE, 
                                       end_color=ExcelColors.HEADER_BLUE, fill_type="solid")
            elif col_idx <= 9:  # 1st Shift columns (6:00am - 6:00pm)
                cell.fill = PatternFill(start_color=ExcelColors.SHIFT1_HEADER, 
                                       end_color=ExcelColors.SHIFT1_HEADER, fill_type="solid")
            elif col_idx <= 12:  # 2nd Shift columns (6:00pm - 6:00am)
                cell.fill = PatternFill(start_color=ExcelColors.SHIFT2_HEADER, 
                                       end_color=ExcelColors.SHIFT2_HEADER, fill_type="solid")
            else:  # Total, Status
                cell.fill = PatternFill(start_color=ExcelColors.INFO, 
                                       end_color=ExcelColors.INFO, fill_type="solid")
            
            cell.border = Border(
                left=Side(style='thin', color=ExcelColors.BORDER_COLOR),
                right=Side(style='thin', color=ExcelColors.BORDER_COLOR),
                top=Side(style='thin', color=ExcelColors.BORDER_COLOR),
                bottom=Side(style='medium', color=ExcelColors.WHITE)
            )
        
        ws.row_dimensions[current_row].height = 30
        current_row += 1
        
        # ===== DATA ROWS =====
        # Group stations by LRU name
        lru_groups = defaultdict(list)
        
        for station_name, station_data in stations.items():
            if " - " in station_name:
                lru_name = station_name.split(" - ")[0].strip()
                location = station_name.split(" - ", 1)[1].strip()
            else:
                lru_name = station_name
                location = ""
            
            lru_groups[lru_name].append({
                'station_name': station_name,
                'location': location,
                'data': station_data
            })
        
        # Add data rows
        for lru_name in sorted(lru_groups.keys()):
            stations_list = lru_groups[lru_name]
            
            for station_info in stations_list:
                station_name = station_info['station_name']
                location = station_info['location']
                data = station_info['data']
                
                # Test description with current status
                test_desc = data.test_description or \
                           f"Current: {data.current} (Min: {data.min_lru}, Max: {data.max_lru})"
                
                # LRU Name
                lru_cell = ws.cell(current_row, 1)
                lru_cell.value = lru_name
                lru_cell.font = Font(size=11, bold=True, name='Calibri')
                lru_cell.alignment = Alignment(horizontal='left', vertical='center')
                lru_cell.fill = PatternFill(start_color=ExcelColors.LIGHT_BLUE, 
                                           end_color=ExcelColors.LIGHT_BLUE, fill_type="solid")
                
                # Test Description
                test_cell = ws.cell(current_row, 2)
                test_cell.value = test_desc
                test_cell.font = Font(size=10, name='Calibri')
                test_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Rack Location
                loc_cell = ws.cell(current_row, 3)
                loc_cell.value = location
                loc_cell.font = Font(size=10, name='Calibri')
                loc_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Get time slot data
                time_slot_data = self._get_time_slot_data(data)
                
                # Add time slot columns with conditional formatting
                total_batches = 0
                for col_idx, time_slot in enumerate(ALL_TIME_SLOTS, 4):
                    cell = ws.cell(current_row, col_idx)
                    value = time_slot_data.get(time_slot, '')
                    cell.value = value if value else ''
                    cell.font = Font(size=11, name='Calibri')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.number_format = '0'  # Whole numbers only
                    
                    # Light background for 1st shift, slightly darker for 2nd shift
                    if col_idx <= 9:  # 1st Shift
                        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type="solid")
                    else:  # 2nd Shift
                        cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type="solid")
                    
                    # Add to total if numeric
                    try:
                        if value:
                            total_batches += int(value)
                    except:
                        pass
                
                # Total column with formula
                total_col = len(ALL_TIME_SLOTS) + 4
                total_cell = ws.cell(current_row, total_col)
                # Create SUM formula for the time slot columns
                start_col = get_column_letter(4)
                end_col = get_column_letter(total_col - 1)
                total_cell.value = f"=SUM({start_col}{current_row}:{end_col}{current_row})"
                total_cell.font = Font(size=11, bold=True, name='Calibri')
                total_cell.alignment = Alignment(horizontal='center', vertical='center')
                total_cell.fill = PatternFill(start_color=ExcelColors.LIGHT_GREEN, 
                                             end_color=ExcelColors.LIGHT_GREEN, fill_type="solid")
                
                # Status column (based on current vs min/max)
                status_cell = ws.cell(current_row, total_col + 1)
                current_status = data.get_status()
                status_cell.value = current_status
                status_cell.font = Font(size=10, bold=True, color=ExcelColors.WHITE, name='Calibri')
                status_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Color code status
                if current_status == "Critical":
                    status_cell.fill = PatternFill(start_color=ExcelColors.CRITICAL, 
                                                  end_color=ExcelColors.CRITICAL, fill_type="solid")
                elif current_status == "Warning":
                    status_cell.fill = PatternFill(start_color=ExcelColors.WARNING, 
                                                  end_color=ExcelColors.WARNING, fill_type="solid")
                else:  # Good
                    status_cell.fill = PatternFill(start_color=ExcelColors.SUCCESS, 
                                                  end_color=ExcelColors.SUCCESS, fill_type="solid")
                
                # Apply borders to all cells in row
                for col_idx in range(1, total_col + 2):
                    cell = ws.cell(current_row, col_idx)
                    cell.border = Border(
                        left=Side(style='thin', color=ExcelColors.BORDER_COLOR),
                        right=Side(style='thin', color=ExcelColors.BORDER_COLOR),
                        top=Side(style='thin', color=ExcelColors.BORDER_COLOR),
                        bottom=Side(style='thin', color=ExcelColors.BORDER_COLOR)
                    )
                
                ws.row_dimensions[current_row].height = 25
                current_row += 1
            
            # Add spacing row after each LRU group
            current_row += 1
        
        # ===== SUMMARY ROW =====
        ws.merge_cells(f'A{current_row}:C{current_row}')
        summary_cell = ws[f'A{current_row}']
        summary_cell.value = "📊 Total Batches per Time Slot:"
        summary_cell.font = Font(size=11, bold=True, color=ExcelColors.TEXT_DARK, name='Calibri')
        summary_cell.alignment = Alignment(horizontal='right', vertical='center')
        summary_cell.fill = PatternFill(start_color=ExcelColors.INFO, end_color=ExcelColors.INFO, fill_type="solid")
        
        # Calculate column totals
        for col_idx in range(4, len(ALL_TIME_SLOTS) + 5):  # +4 for initial columns, +1 for Total column
            cell = ws.cell(current_row, col_idx)
            col_letter = get_column_letter(col_idx)
            # Sum from row 5 (first data row) to current_row - 1
            cell.value = f"=SUM({col_letter}5:{col_letter}{current_row - 1})"
            cell.font = Font(size=11, bold=True, color=ExcelColors.WHITE, name='Calibri')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color=ExcelColors.INFO, end_color=ExcelColors.INFO, fill_type="solid")
            cell.border = Border(
                left=Side(style='thin', color=ExcelColors.BORDER_COLOR),
                right=Side(style='thin', color=ExcelColors.BORDER_COLOR),
                top=Side(style='medium', color=ExcelColors.TEXT_DARK),
                bottom=Side(style='medium', color=ExcelColors.TEXT_DARK)
            )
        
        ws.row_dimensions[current_row].height = 30
        
        # ===== COLUMN WIDTHS =====
        ws.column_dimensions['A'].width = 25  # LRU
        ws.column_dimensions['B'].width = 40  # Test to Schedule
        ws.column_dimensions['C'].width = 18  # Rack Location
        
        # Time slot columns
        for col_idx in range(4, len(ALL_TIME_SLOTS) + 4):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
        
        ws.column_dimensions[get_column_letter(len(ALL_TIME_SLOTS) + 4)].width = 12  # Total
        ws.column_dimensions[get_column_letter(len(ALL_TIME_SLOTS) + 5)].width = 12  # Status
        
        # ===== FREEZE PANES =====
        ws.freeze_panes = 'D5'  # Freeze header rows and first 3 columns
        
        # ===== PRINT SETTINGS =====
        ws.print_options.horizontalCentered = True
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # Allow multiple pages vertically if needed
        
        wb.save(filename)
        logger.info(f"Exported professional FC schedule to Excel: {filename}")
    
    def _export_to_csv_legacy(self, filename: str, stations: Dict[str, Station]) -> None:
        """Export stations in legacy CSV format (kept for backward compatibility)."""
        # Group stations by LRU name
        lru_groups = defaultdict(list)
        
        for station_name, station_data in stations.items():
            if " - " in station_name:
                lru_name = station_name.split(" - ")[0].strip()
                location = station_name.split(" - ", 1)[1].strip()
            else:
                lru_name = station_name
                location = ""
            
            lru_groups[lru_name].append({
                'station_name': station_name,
                'location': location,
                'data': station_data
            })
        
        # Prepare CSV data
        csv_data = []
        
        # Headers
        header1 = ['', '', '', '1st Shift - Record # of Batches to schedule', '', '', '', '', '', 
                  '2nd Shift - Record # of Batches to schedule', '', '']
        csv_data.append(header1)
        
        header2 = ['LRU', 'Test to Schedule', 'Rack Location'] + ALL_TIME_SLOTS
        csv_data.append(header2)
        
        # Data
        for lru_name in sorted(lru_groups.keys()):
            stations_list = lru_groups[lru_name]
            
            for station_info in stations_list:
                station_name = station_info['station_name']
                location = station_info['location']
                data = station_info['data']
                
                test_desc = data.test_description or f"Current: {data.current} (Min: {data.min_lru}, Max: {data.max_lru})"
                
                row = [lru_name, test_desc, location]
                
                # Add time slot data
                time_slot_data = self._get_time_slot_data(data)
                for time_slot in ALL_TIME_SLOTS:
                    row.append(time_slot_data.get(time_slot, ''))
                
                csv_data.append(row)
            
            csv_data.append([''] * len(header2))
        
        # Write CSV
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(csv_data)
        
        logger.info(f"Exported FC schedule to: {filename}")
    
    def _get_time_slot_data(self, station: Station) -> Dict[str, str]:
        """Extract time slot data from station history."""
        result = {}
        
        for time_slot, (start_hour, end_hour) in TIME_SLOT_MAP.items():
            latest_value = None
            
            for entry in station.history:
                try:
                    dt = datetime.strptime(entry.timestamp, TIMESTAMP_FORMAT)
                    hour = dt.hour
                    
                    if start_hour <= hour < end_hour:
                        latest_value = entry.count
                except:
                    continue
            
            if latest_value is not None:
                result[time_slot] = str(latest_value)
        
        return result
