import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
from datetime import datetime
from typing import Optional, Dict, List, Any
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import threading

# Application version
APP_VERSION = "1.0.0"
UPDATE_CHECK_URL = "https://raw.githubusercontent.com/HaltTheGrey/lru-tracker/main/version.json"

class LRUTrackerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("FC LRU Pull System Tracker")
        self.root.geometry("1400x800")
        self.root.configure(bg='#f0f0f0')
        
        self.stations = {}
        self.history = []
        self.data_file = "lru_data.json"
        
        self.load_data()
        self.create_ui()
        self.refresh_display()
    
    def create_ui(self):
        # Title
        title_frame = tk.Frame(self.root, bg='#2c3e50', height=60)
        title_frame.pack(fill='x', pady=(0, 5))
        title_frame.pack_propagate(False)
        
        tk.Label(title_frame, text="🏭 FC LRU Pull System Tracker", 
                font=('Arial', 20, 'bold'), bg='#2c3e50', fg='white').pack(pady=15)
        
        # Main container
        main_container = tk.Frame(self.root, bg='#f0f0f0')
        main_container.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Left panel - Station Management
        left_panel = tk.Frame(main_container, bg='white', relief='raised', bd=2)
        left_panel.pack(side='left', fill='both', expand=True, padx=(0, 5))
        
        # Station controls
        control_frame = tk.Frame(left_panel, bg='white', pady=5)
        control_frame.pack(fill='x', padx=5)
        
        tk.Label(control_frame, text="Station Management", 
                font=('Arial', 13, 'bold'), bg='white').pack(anchor='w', pady=(0, 5))
        
        btn_frame = tk.Frame(control_frame, bg='white')
        btn_frame.pack(fill='x')
        
        tk.Button(btn_frame, text="➕ Add", command=self.add_station_dialog,
                 bg='#27ae60', fg='white', font=('Arial', 9, 'bold'),
                 padx=10, pady=5, cursor='hand2').pack(side='left', padx=2)
        
        tk.Button(btn_frame, text="✏️ Edit", command=self.edit_station_dialog,
                 bg='#3498db', fg='white', font=('Arial', 9, 'bold'),
                 padx=10, pady=5, cursor='hand2').pack(side='left', padx=2)
        
        tk.Button(btn_frame, text="🗑️ Delete", command=self.delete_station,
                 bg='#e74c3c', fg='white', font=('Arial', 9, 'bold'),
                 padx=10, pady=5, cursor='hand2').pack(side='left', padx=2)
        
        # Stations tree view
        tree_frame = tk.Frame(left_panel, bg='white')
        tree_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Scrollbars
        v_scroll = ttk.Scrollbar(tree_frame, orient='vertical')
        h_scroll = ttk.Scrollbar(tree_frame, orient='horizontal')
        
        self.tree = ttk.Treeview(tree_frame, 
                                columns=('Current', 'Min', 'Max', 'Status'),
                                show='tree headings',
                                yscrollcommand=v_scroll.set,
                                xscrollcommand=h_scroll.set,
                                height=20)
        
        v_scroll.config(command=self.tree.yview)
        h_scroll.config(command=self.tree.xview)
        
        self.tree.heading('#0', text='Station Name')
        self.tree.heading('Current', text='Current LRU')
        self.tree.heading('Min', text='Min')
        self.tree.heading('Max', text='Max')
        self.tree.heading('Status', text='Status')
        
        self.tree.column('#0', width=200)
        self.tree.column('Current', width=100, anchor='center')
        self.tree.column('Min', width=80, anchor='center')
        self.tree.column('Max', width=80, anchor='center')
        self.tree.column('Status', width=120, anchor='center')
        
        # Bind selection event to auto-fill update box
        self.tree.bind('<<TreeviewSelect>>', self.on_station_select)
        
        self.tree.pack(side='left', fill='both', expand=True)
        v_scroll.pack(side='right', fill='y')
        h_scroll.pack(side='bottom', fill='x')
        
        # Configure tags for color coding
        self.tree.tag_configure('under_min', background='#ffcccc')
        self.tree.tag_configure('at_max', background='#fff4cc')
        self.tree.tag_configure('normal', background='#ccffcc')
        
        # Right panel - Update & Export (scrollable)
        right_panel = tk.Frame(main_container, bg='white', relief='raised', bd=2, width=380)
        right_panel.pack(side='right', fill='both', padx=(5, 0))
        right_panel.pack_propagate(False)
        
        # Create canvas for scrolling
        canvas = tk.Canvas(right_panel, bg='white', highlightthickness=0)
        scrollbar = ttk.Scrollbar(right_panel, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Enable mousewheel scrolling
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        # Update LRU section
        update_frame = tk.LabelFrame(scrollable_frame, text="Update LRU Count", 
                                    font=('Arial', 12, 'bold'), bg='white', padx=10, pady=10)
        update_frame.pack(fill='x', padx=5, pady=5)
        
        tk.Label(update_frame, text="Station:", bg='white', font=('Arial', 10)).pack(anchor='w')
        self.update_station_var = tk.StringVar()
        self.update_station_combo = ttk.Combobox(update_frame, textvariable=self.update_station_var,
                                                 state='readonly', font=('Arial', 10))
        self.update_station_combo.pack(fill='x', pady=(3, 8))
        
        tk.Label(update_frame, text="New LRU Count:", bg='white', font=('Arial', 10)).pack(anchor='w')
        self.update_count_var = tk.StringVar()
        tk.Entry(update_frame, textvariable=self.update_count_var, 
                font=('Arial', 11), justify='center').pack(fill='x', pady=(3, 8))
        
        tk.Button(update_frame, text="📝 Update Count", command=self.update_lru_count,
                 bg='#3498db', fg='white', font=('Arial', 10, 'bold'),
                 padx=15, pady=8, cursor='hand2').pack(fill='x')
        
        # Statistics section
        stats_frame = tk.LabelFrame(scrollable_frame, text="Statistics", 
                                   font=('Arial', 12, 'bold'), bg='white', padx=10, pady=8)
        stats_frame.pack(fill='x', padx=5, pady=5)
        
        self.stats_label = tk.Label(stats_frame, text="", bg='white', 
                                   font=('Arial', 9), justify='left', anchor='w')
        self.stats_label.pack(fill='x')
        
        # Export section
        export_frame = tk.LabelFrame(scrollable_frame, text="Export & Reports", 
                                    font=('Arial', 12, 'bold'), bg='white', padx=10, pady=8)
        export_frame.pack(fill='x', padx=5, pady=5)
        
        tk.Button(export_frame, text="📊 New Excel Report", 
                 command=self.export_new_report,
                 bg='#27ae60', fg='white', font=('Arial', 9, 'bold'),
                 padx=10, pady=6, cursor='hand2').pack(fill='x', pady=3)
        
        tk.Button(export_frame, text="📈 Append to Existing", 
                 command=self.append_to_existing,
                 bg='#16a085', fg='white', font=('Arial', 9, 'bold'),
                 padx=10, pady=6, cursor='hand2').pack(fill='x', pady=3)
        
        tk.Button(export_frame, text="📉 View Trends", 
                 command=self.view_trends,
                 bg='#2980b9', fg='white', font=('Arial', 9, 'bold'),
                 padx=10, pady=6, cursor='hand2').pack(fill='x', pady=3)
        
        tk.Button(export_frame, text="📅 Export FC Schedule Format", 
                 command=self.export_fc_schedule_format,
                 bg='#d35400', fg='white', font=('Arial', 9, 'bold'),
                 padx=10, pady=6, cursor='hand2').pack(fill='x', pady=3)
        
        # Template section
        template_frame = tk.LabelFrame(scrollable_frame, text="Template Import/Export", 
                                      font=('Arial', 12, 'bold'), bg='white', padx=10, pady=8)
        template_frame.pack(fill='x', padx=5, pady=5)
        
        tk.Button(template_frame, text="📥 Download Template", 
                 command=self.download_template,
                 bg='#8e44ad', fg='white', font=('Arial', 9, 'bold'),
                 padx=10, pady=6, cursor='hand2').pack(fill='x', pady=3)
        
        tk.Button(template_frame, text="📤 Import from Template", 
                 command=self.import_from_template,
                 bg='#9b59b6', fg='white', font=('Arial', 9, 'bold'),
                 padx=10, pady=6, cursor='hand2').pack(fill='x', pady=3)
        
        tk.Button(template_frame, text="📋 Import FC Schedule CSV", 
                 command=self.import_fc_schedule,
                 bg='#e67e22', fg='white', font=('Arial', 9, 'bold'),
                 padx=10, pady=6, cursor='hand2').pack(fill='x', pady=3)
        
        # Updates section
        updates_frame = tk.LabelFrame(scrollable_frame, text="Application", 
                                      font=('Arial', 12, 'bold'), bg='white', padx=10, pady=8)
        updates_frame.pack(fill='x', padx=5, pady=5)
        
        tk.Button(updates_frame, text="🔄 Check for Updates", 
                 command=self.check_for_updates,
                 bg='#27ae60', fg='white', font=('Arial', 9, 'bold'),
                 padx=10, pady=6, cursor='hand2').pack(fill='x', pady=3)
        
        # Version label
        version_label = tk.Label(scrollable_frame, text=f"Version {APP_VERSION}", 
                                font=('Arial', 8), bg='#ecf0f1', fg='#7f8c8d')
        version_label.pack(pady=(5, 10))
        
        # Save button
        save_frame = tk.Frame(scrollable_frame, bg='white')
        save_frame.pack(fill='x', padx=5, pady=10)
        
        tk.Button(save_frame, text="💾 Save Data", command=self.save_data,
                 bg='#95a5a6', fg='white', font=('Arial', 9, 'bold'),
                 padx=10, pady=6, cursor='hand2').pack(fill='x')
    
    def add_station_dialog(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Add New Station")
        dialog.geometry("400x300")
        dialog.configure(bg='white')
        dialog.transient(self.root)
        dialog.grab_set()
        
        tk.Label(dialog, text="Add New Station", font=('Arial', 16, 'bold'), 
                bg='white').pack(pady=15)
        
        frame = tk.Frame(dialog, bg='white', padx=20)
        frame.pack(fill='both', expand=True)
        
        tk.Label(frame, text="Station Name:", bg='white', font=('Arial', 11)).pack(anchor='w', pady=(10, 5))
        name_var = tk.StringVar()
        tk.Entry(frame, textvariable=name_var, font=('Arial', 11)).pack(fill='x', pady=(0, 15))
        
        tk.Label(frame, text="Minimum LRU:", bg='white', font=('Arial', 11)).pack(anchor='w', pady=(0, 5))
        min_var = tk.StringVar(value="5")
        tk.Entry(frame, textvariable=min_var, font=('Arial', 11)).pack(fill='x', pady=(0, 15))
        
        tk.Label(frame, text="Maximum LRU:", bg='white', font=('Arial', 11)).pack(anchor='w', pady=(0, 5))
        max_var = tk.StringVar(value="20")
        tk.Entry(frame, textvariable=max_var, font=('Arial', 11)).pack(fill='x', pady=(0, 15))
        
        def save_station():
            name = name_var.get().strip()
            
            # Validate station name
            if not self.validate_station_name(name):
                messagebox.showerror("Error", 
                    "Invalid station name!\n\n"
                    "Station names must:\n"
                    "• Not be empty\n"
                    "• Be less than 200 characters\n"
                    "• Contain only letters, numbers, spaces, and basic punctuation")
                return
            
            # Validate and parse min value
            min_valid, min_val = self.validate_number(min_var.get(), 0, 999999)
            if not min_valid:
                messagebox.showerror("Error", 
                    "Invalid minimum value!\n"
                    "Must be a number between 0 and 999,999")
                return
            
            # Validate and parse max value
            max_valid, max_val = self.validate_number(max_var.get(), 0, 999999)
            if not max_valid:
                messagebox.showerror("Error", 
                    "Invalid maximum value!\n"
                    "Must be a number between 0 and 999,999")
                return
            
            if name in self.stations:
                messagebox.showerror("Error", "Station already exists!")
                return
            
            if min_val > max_val:
                messagebox.showerror("Error", "Minimum cannot be greater than Maximum!")
                return
            
            self.stations[name] = {
                'current': 0,
                'min': min_val,
                'max': max_val,
                'history': []
            }
            
            self.save_data()
            self.refresh_display()
            dialog.destroy()
            messagebox.showinfo("Success", f"Station '{name}' added successfully!")
        
        btn_frame = tk.Frame(dialog, bg='white')
        btn_frame.pack(fill='x', padx=20, pady=15)
        
        tk.Button(btn_frame, text="Add Station", command=save_station,
                 bg='#27ae60', fg='white', font=('Arial', 11, 'bold'),
                 padx=20, pady=8).pack(side='left', padx=5)
        
        tk.Button(btn_frame, text="Cancel", command=dialog.destroy,
                 bg='#95a5a6', fg='white', font=('Arial', 11, 'bold'),
                 padx=20, pady=8).pack(side='right', padx=5)
    
    def edit_station_dialog(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a station to edit!")
            return
        
        station_name = self.tree.item(selected[0])['text']
        station = self.stations[station_name]
        
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Edit Station: {station_name}")
        dialog.geometry("400x250")
        dialog.configure(bg='white')
        dialog.transient(self.root)
        dialog.grab_set()
        
        tk.Label(dialog, text=f"Edit Station: {station_name}", 
                font=('Arial', 16, 'bold'), bg='white').pack(pady=15)
        
        frame = tk.Frame(dialog, bg='white', padx=20)
        frame.pack(fill='both', expand=True)
        
        tk.Label(frame, text="Minimum LRU:", bg='white', font=('Arial', 11)).pack(anchor='w', pady=(0, 5))
        min_var = tk.StringVar(value=str(station['min']))
        tk.Entry(frame, textvariable=min_var, font=('Arial', 11)).pack(fill='x', pady=(0, 15))
        
        tk.Label(frame, text="Maximum LRU:", bg='white', font=('Arial', 11)).pack(anchor='w', pady=(0, 5))
        max_var = tk.StringVar(value=str(station['max']))
        tk.Entry(frame, textvariable=max_var, font=('Arial', 11)).pack(fill='x', pady=(0, 15))
        
        def save_changes():
            try:
                min_val = int(min_var.get())
                max_val = int(max_var.get())
                
                if min_val < 0 or max_val < 0:
                    messagebox.showerror("Error", "Min and Max must be positive!")
                    return
                
                if min_val > max_val:
                    messagebox.showerror("Error", "Min cannot be greater than Max!")
                    return
                
                self.stations[station_name]['min'] = min_val
                self.stations[station_name]['max'] = max_val
                
                self.save_data()
                self.refresh_display()
                dialog.destroy()
                messagebox.showinfo("Success", "Station updated successfully!")
                
            except ValueError:
                messagebox.showerror("Error", "Min and Max must be valid numbers!")
        
        btn_frame = tk.Frame(dialog, bg='white')
        btn_frame.pack(fill='x', padx=20, pady=15)
        
        tk.Button(btn_frame, text="Save Changes", command=save_changes,
                 bg='#3498db', fg='white', font=('Arial', 11, 'bold'),
                 padx=20, pady=8).pack(side='left', padx=5)
        
        tk.Button(btn_frame, text="Cancel", command=dialog.destroy,
                 bg='#95a5a6', fg='white', font=('Arial', 11, 'bold'),
                 padx=20, pady=8).pack(side='right', padx=5)
    
    def delete_station(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a station to delete!")
            return
        
        station_name = self.tree.item(selected[0])['text']
        
        if messagebox.askyesno("Confirm Delete", 
                              f"Are you sure you want to delete '{station_name}'?\nAll history will be lost."):
            del self.stations[station_name]
            self.save_data()
            self.refresh_display()
            messagebox.showinfo("Success", f"Station '{station_name}' deleted!")
    
    def update_lru_count(self):
        station_name = self.update_station_var.get()
        if not station_name:
            messagebox.showwarning("Warning", "Please select a station!")
            return
        
        try:
            new_count = int(self.update_count_var.get())
            if new_count < 0:
                messagebox.showerror("Error", "LRU count cannot be negative!")
                return
            
            # Update current count
            self.stations[station_name]['current'] = new_count
            
            # Add to history
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.stations[station_name]['history'].append({
                'timestamp': timestamp,
                'count': new_count
            })
            
            # Add to global history
            self.history.append({
                'station': station_name,
                'timestamp': timestamp,
                'count': new_count,
                'min': self.stations[station_name]['min'],
                'max': self.stations[station_name]['max']
            })
            
            self.save_data()
            self.refresh_display()
            self.update_count_var.set("")
            
            messagebox.showinfo("Success", f"Updated '{station_name}' to {new_count} LRUs!")
            
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid number!")
    
    def on_station_select(self, event):
        """Auto-fill the update section when a station is selected"""
        selected = self.tree.selection()
        if selected:
            station_name = self.tree.item(selected[0])['text']
            # Set the dropdown to the selected station
            self.update_station_var.set(station_name)
            # Focus on the count entry field for quick input
            # Find the entry widget and focus it
            for widget in self.root.winfo_children():
                self._focus_count_entry(widget)
    
    def _focus_count_entry(self, widget):
        """Recursively find and focus the count entry widget"""
        if isinstance(widget, tk.Entry) and widget['textvariable'] == str(self.update_count_var):
            widget.focus_set()
            widget.select_range(0, tk.END)
            return True
        for child in widget.winfo_children():
            if self._focus_count_entry(child):
                return True
        return False
    
    def refresh_display(self):
        # Clear tree
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Populate tree
        under_min_count = 0
        at_max_count = 0
        normal_count = 0
        
        for name, data in sorted(self.stations.items()):
            current = data['current']
            min_val = data['min']
            max_val = data['max']
            
            if current < min_val:
                status = "⚠️ UNDER MIN"
                tag = 'under_min'
                under_min_count += 1
            elif current >= max_val:
                status = "🔴 AT/OVER MAX"
                tag = 'at_max'
                at_max_count += 1
            else:
                status = "✅ Normal"
                tag = 'normal'
                normal_count += 1
            
            self.tree.insert('', 'end', text=name, 
                           values=(current, min_val, max_val, status),
                           tags=(tag,))
        
        # Update combo box
        station_names = list(self.stations.keys())
        self.update_station_combo['values'] = station_names
        if station_names and not self.update_station_var.get():
            self.update_station_var.set(station_names[0])
        
        # Update statistics
        total_stations = len(self.stations)
        stats_text = f"""
Total Stations: {total_stations}

✅ Normal: {normal_count}
⚠️  Under Min: {under_min_count}
🔴 At/Over Max: {at_max_count}
        """
        self.stats_label.config(text=stats_text.strip())
    
    def export_new_report(self):
        if not self.stations:
            messagebox.showwarning("Warning", "No stations to export!")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"LRU_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws: Worksheet = wb.active  # type: ignore
            ws.title = "Current Status"
            
            # Header
            headers = ["Station Name", "Current LRU", "Min", "Max", "Status", "Last Updated"]
            ws.append(headers)
            
            # Style header
            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            
            for col in range(1, len(headers) + 1):
                cell = ws.cell(1, col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data
            for name, data in sorted(self.stations.items()):
                current = data['current']
                min_val = data['min']
                max_val = data['max']
                
                if current < min_val:
                    status = "UNDER MIN"
                    status_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif current >= max_val:
                    status = "AT/OVER MAX"
                    status_fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
                else:
                    status = "Normal"
                    status_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                
                last_updated = data['history'][-1]['timestamp'] if data['history'] else "Never"
                
                row = [name, current, min_val, max_val, status, last_updated]
                ws.append(row)
                
                # Color code status
                status_cell = ws.cell(ws.max_row, 5)
                status_cell.fill = status_fill
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 20
            
            # Create history sheet if there's history
            if self.history:
                ws_history = wb.create_sheet("History")
                history_headers = ["Station", "Timestamp", "Count", "Min", "Max"]
                ws_history.append(history_headers)
                
                # Style header
                for col in range(1, len(history_headers) + 1):
                    cell = ws_history.cell(1, col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Add history data
                for record in self.history:
                    ws_history.append([
                        record['station'],
                        record['timestamp'],
                        record['count'],
                        record['min'],
                        record['max']
                    ])
                
                # Adjust column widths
                ws_history.column_dimensions['A'].width = 25
                ws_history.column_dimensions['B'].width = 20
                ws_history.column_dimensions['C'].width = 12
                ws_history.column_dimensions['D'].width = 10
                ws_history.column_dimensions['E'].width = 10
            
            wb.save(filename)
            messagebox.showinfo("Success", f"Report exported successfully to:\n{filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export report:\n{str(e)}")
    
    def append_to_existing(self):
        if not self.stations:
            messagebox.showwarning("Warning", "No stations to export!")
            return
        
        filename = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.load_workbook(filename)
            
            # Create new sheet with timestamp
            sheet_name = f"Snapshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            ws = wb.create_sheet(sheet_name)
            
            # Header
            headers = ["Station Name", "Current LRU", "Min", "Max", "Status", "Timestamp"]
            ws.append(headers)
            
            # Style header
            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            
            for col in range(1, len(headers) + 1):
                cell = ws.cell(1, col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for name, data in sorted(self.stations.items()):
                current = data['current']
                min_val = data['min']
                max_val = data['max']
                
                if current < min_val:
                    status = "UNDER MIN"
                    status_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif current >= max_val:
                    status = "AT/OVER MAX"
                    status_fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
                else:
                    status = "Normal"
                    status_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                
                row = [name, current, min_val, max_val, status, timestamp]
                ws.append(row)
                
                status_cell = ws.cell(ws.max_row, 5)
                status_cell.fill = status_fill
            
            # Adjust column widths
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws.column_dimensions[col].width = 18
            
            wb.save(filename)
            messagebox.showinfo("Success", f"Data appended to existing report!\nNew sheet: {sheet_name}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to append to report:\n{str(e)}")
    
    def view_trends(self):
        if not self.stations:
            messagebox.showwarning("Warning", "No stations available!")
            return
        
        # Create selection dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("View Trends")
        dialog.geometry("400x200")
        dialog.configure(bg='white')
        dialog.transient(self.root)
        dialog.grab_set()
        
        tk.Label(dialog, text="Select Station for Trend Analysis", 
                font=('Arial', 14, 'bold'), bg='white').pack(pady=15)
        
        station_var = tk.StringVar()
        station_combo = ttk.Combobox(dialog, textvariable=station_var,
                                    values=list(self.stations.keys()),
                                    state='readonly', font=('Arial', 11))
        station_combo.pack(padx=20, pady=10, fill='x')
        station_combo.current(0)
        
        def generate_trend():
            station_name = station_var.get()
            if not station_name:
                messagebox.showwarning("Warning", "Please select a station!")
                return
            
            history = self.stations[station_name]['history']
            if not history:
                messagebox.showinfo("Info", f"No history data available for '{station_name}'")
                return
            
            # Create Excel with chart
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"Trend_{station_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            )
            
            if not filename:
                return
            
            try:
                wb = openpyxl.Workbook()
                ws: Worksheet = wb.active  # type: ignore
                ws.title = f"{station_name} Trends"
                
                # Headers
                ws.append(["Timestamp", "LRU Count", "Min", "Max"])
                
                min_val = self.stations[station_name]['min']
                max_val = self.stations[station_name]['max']
                
                # Data
                for record in history:
                    ws.append([record['timestamp'], record['count'], min_val, max_val])
                
                # Create chart
                chart = LineChart()
                chart.title = f"LRU Trend for {station_name}"
                chart.style = 10
                chart.x_axis.title = "Time"
                chart.y_axis.title = "LRU Count"
                chart.height = 15
                chart.width = 25
                
                data = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=len(history) + 1)
                cats = Reference(ws, min_col=1, min_row=2, max_row=len(history) + 1)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                
                ws.add_chart(chart, "F2")
                
                wb.save(filename)
                dialog.destroy()
                messagebox.showinfo("Success", f"Trend report created:\n{filename}")
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create trend report:\n{str(e)}")
        
        tk.Button(dialog, text="Generate Trend Report", command=generate_trend,
                 bg='#2980b9', fg='white', font=('Arial', 11, 'bold'),
                 padx=20, pady=10).pack(pady=20)
    
    def export_fc_schedule_format(self):
        """Export in FC Standard Work Spreadsheet format with time tracking"""
        if not self.stations:
            messagebox.showwarning("Warning", "No stations to export!")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"FC_Schedule_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
        
        if not filename:
            return
        
        try:
            import csv
            from collections import defaultdict
            
            # Group stations by LRU name (extract base LRU from "LRU - Location" format)
            lru_groups = defaultdict(list)
            
            for station_name, station_data in self.stations.items():
                # Try to extract LRU name (before the " - ")
                if " - " in station_name:
                    lru_name = station_name.split(" - ")[0].strip()
                    location = station_name.split(" - ", 1)[1].strip()
                else:
                    lru_name = station_name
                    location = ""
                
                lru_groups[lru_name].append({
                    'station_name': station_name,
                    'location': location,
                    'data': station_data
                })
            
            # Time slots for the schedule (matching your format)
            time_slots_1st = ['6AM', '8AM', '10AM', '12PM', '2PM', '4PM']
            time_slots_2nd = ['6PM', '8PM', '10PM', '12AM']
            all_time_slots = time_slots_1st + time_slots_2nd
            
            # Prepare data for CSV
            csv_data = []
            
            # Header row 1
            header1 = ['', '', '', '1st Shift - Record # of Batches to schedule', '', '', '', '', '', 
                      '2nd Shift - Record # of Batches to schedule', '', '']
            csv_data.append(header1)
            
            # Header row 2
            header2 = ['LRU', 'Test to Schedule (Yellow = FC Moves Hardware)', 'Rack Location'] + all_time_slots
            csv_data.append(header2)
            
            # Add data for each LRU group
            for lru_name in sorted(lru_groups.keys()):
                stations = lru_groups[lru_name]
                
                for station_info in stations:
                    station_name = station_info['station_name']
                    location = station_info['location']
                    data = station_info['data']
                    
                    # Create test description
                    current = data['current']
                    min_val = data['min']
                    max_val = data['max']
                    test_desc = data.get('test_description', f"Current: {current} (Min: {min_val}, Max: {max_val})")
                    
                    # Create row with LRU, test description, location
                    row = [lru_name, test_desc, location]
                    
                    # Add time slot data based on history
                    time_slot_data = self._get_time_slot_data(station_name, data)
                    
                    for time_slot in all_time_slots:
                        row.append(time_slot_data.get(time_slot, ''))
                    
                    csv_data.append(row)
                
                # Add empty row between LRU groups
                csv_data.append([''] * len(header2))
            
            # Write to CSV
            if filename.endswith('.csv'):
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerows(csv_data)
                messagebox.showinfo("Success", f"FC Schedule exported to:\n{filename}")
            
            elif filename.endswith('.xlsx'):
                # Export to Excel with formatting
                wb = openpyxl.Workbook()
                ws: Worksheet = wb.active  # type: ignore
                ws.title = "FC Schedule"
                
                # Write data
                for row_idx, row_data in enumerate(csv_data, 1):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row_idx, col_idx)
                        cell.value = value
                        
                        # Format headers
                        if row_idx == 1:
                            cell.font = Font(bold=True, size=11)
                            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                        elif row_idx == 2:
                            cell.font = Font(bold=True, size=10)
                            cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Add borders
                        if row_idx >= 2:
                            cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                
                # Set column widths
                ws.column_dimensions['A'].width = 15  # LRU
                ws.column_dimensions['B'].width = 40  # Test Description
                ws.column_dimensions['C'].width = 20  # Rack Location
                
                # Time slot columns
                for col in range(4, 4 + len(all_time_slots)):
                    ws.column_dimensions[get_column_letter(col)].width = 10
                
                # Merge header cells for shift labels
                ws.merge_cells('D1:I1')  # 1st Shift
                ws.merge_cells('J1:M1')  # 2nd Shift
                
                # Center align merged cells
                ws['D1'].alignment = Alignment(horizontal='center', vertical='center')
                ws['J1'].alignment = Alignment(horizontal='center', vertical='center')
                
                wb.save(filename)
                messagebox.showinfo("Success", f"FC Schedule exported to:\n{filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export FC schedule:\n{str(e)}")
    
    def _get_time_slot_data(self, station_name, station_data):
        """Extract time slot data from history for FC schedule format"""
        time_slot_map = {
            '6AM': (6, 8),
            '8AM': (8, 10),
            '10AM': (10, 12),
            '12PM': (12, 14),
            '2PM': (14, 18),
            '4PM': (16, 18),
            '6PM': (18, 20),
            '8PM': (20, 22),
            '10PM': (22, 24),
            '12AM': (0, 2)
        }
        
        result = {}
        history = station_data.get('history', [])
        
        # Group history entries by time slot
        for time_slot, (start_hour, end_hour) in time_slot_map.items():
            count = 0
            latest_value = None
            
            for entry in history:
                timestamp = entry.get('timestamp', '')
                try:
                    # Parse timestamp
                    dt = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
                    hour = dt.hour
                    
                    # Check if this entry falls in the time slot
                    if start_hour <= hour < end_hour:
                        count += 1
                        latest_value = entry.get('count', 0)
                
                except:
                    continue
            
            # Store the latest value if any updates in this time slot
            if latest_value is not None:
                result[time_slot] = str(latest_value)
        
        return result
    
    def save_data(self):
        data = {
            'stations': self.stations,
            'history': self.history
        }
        
        try:
            # Create backup before saving
            if os.path.exists(self.data_file):
                backup_file = self.data_file + '.backup'
                try:
                    import shutil
                    shutil.copy2(self.data_file, backup_file)
                except:
                    pass  # Backup is best-effort
            
            # Write to temporary file first for atomic write
            temp_file = self.data_file + '.tmp'
            with open(temp_file, 'w') as f:
                json.dump(data, f, indent=2)
            
            # Atomic rename
            if os.path.exists(self.data_file):
                os.remove(self.data_file)
            os.rename(temp_file, self.data_file)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save data:\n{str(e)}")
    
    def validate_station_name(self, name: str) -> bool:
        """Validate station name for security and consistency"""
        if not name or not isinstance(name, str):
            return False
        if not name.strip():
            return False
        if len(name) > 200:  # Reasonable limit
            return False
        # Allow alphanumeric, spaces, hyphens, underscores, and common punctuation
        import re
        if not re.match(r'^[\w\s\-.,()#]+$', name):
            return False
        return True
    
    def validate_number(self, value: str, min_val: int = 0, max_val: int = 999999) -> tuple[bool, int]:
        """Validate numeric input and return (is_valid, number)"""
        try:
            num = int(str(value).strip())
            if min_val <= num <= max_val:
                return True, num
            return False, 0
        except (ValueError, TypeError):
            return False, 0
    
    def sanitize_filename(self, filename: str) -> str:
        """Sanitize filename to prevent path traversal attacks"""
        if not filename:
            return "export.xlsx"
        # Get just the basename (removes any path components)
        filename = os.path.basename(filename)
        # Remove potentially dangerous characters
        import re
        filename = re.sub(r'[^\w\s\-.]', '_', filename)
        # Ensure it has an extension
        if not filename.endswith('.xlsx') and not filename.endswith('.csv'):
            filename += '.xlsx'
        return filename
    
    def load_data(self):
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r') as f:
                    content = f.read()
                    # Validate JSON before parsing
                    try:
                        data = json.loads(content)
                    except json.JSONDecodeError as e:
                        messagebox.showerror("Error", 
                            "Data file is corrupted. Starting with empty data.\n"
                            "A backup may exist in lru_data.json.backup")
                        self.stations = {}
                        self.history = []
                        return
                    
                    # Validate data structure
                    if not isinstance(data, dict):
                        raise ValueError("Invalid data format")
                    
                    self.stations = data.get('stations', {})
                    self.history = data.get('history', [])
                    
                    # Validate loaded data
                    if not isinstance(self.stations, dict):
                        self.stations = {}
                    if not isinstance(self.history, list):
                        self.history = []
                    
            except Exception as e:
                messagebox.showerror("Error", 
                    f"Failed to load data. Starting fresh.\n"
                    f"Check lru_data.json.backup if data was lost.")
                self.stations = {}
                self.history = []
    
    def download_template(self):
        """Generate and download an Excel template for bulk station import"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="LRU_Station_Template.xlsx"
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws: Worksheet = wb.active  # type: ignore
            ws.title = "Station Setup"
            
            # Create header with instructions
            ws.merge_cells('A1:E1')
            title_cell = ws['A1']
            title_cell.value = "LRU Station Setup Template"
            title_cell.font = Font(size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30
            
            # Instructions
            ws.merge_cells('A2:E2')
            instructions = ws['A2']
            instructions.value = "Fill in your stations below. Do not modify the header row (row 3)."
            instructions.font = Font(italic=True)
            instructions.alignment = Alignment(horizontal='center')
            ws.row_dimensions[2].height = 25
            
            # Column headers
            headers = ["Station Name", "Min LRU", "Max LRU", "Current LRU (Optional)", "Notes (Optional)"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(3, col_num)
                cell.value = header  # type: ignore
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.row_dimensions[3].height = 25
            
            # Add example rows with explanations
            examples = [
                ["Pack Station 1", 5, 20, 10, "Main packing area"],
                ["Dock Door A", 10, 30, 15, "Receiving dock"],
                ["Induct Station 1", 3, 15, 8, "Induction line"],
                ["Palletize Zone B", 8, 25, 12, "Palletizing area"],
                ["", "", "", "", "← Add more stations below"]
            ]
            
            # Add example data with light background
            example_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
            for row_num, example in enumerate(examples, 4):
                for col_num, value in enumerate(example, 1):
                    cell = ws.cell(row_num, col_num)
                    cell.value = value
                    if row_num < 8:  # Color only example rows
                        cell.fill = example_fill
                    if col_num in [2, 3, 4]:  # Number columns
                        cell.alignment = Alignment(horizontal='center')
            
            # Add empty rows for user to fill
            for row_num in range(9, 30):
                for col_num in range(1, 6):
                    cell = ws.cell(row_num, col_num)
                    cell.value = ""  # type: ignore
            
            # Set column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 30
            
            # Add borders to all cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(3, 30):
                for col in range(1, 6):
                    ws.cell(row, col).border = thin_border
            
            # Create Instructions sheet
            ws_instructions = wb.create_sheet("Instructions")
            
            instruction_text = [
                ["LRU Station Template - Instructions", ""],
                ["", ""],
                ["How to Use This Template:", ""],
                ["", ""],
                ["1. Fill in Station Information", ""],
                ["   • Station Name: Enter a unique name for each station", ""],
                ["   • Min LRU: Minimum threshold (system alerts when below this)", ""],
                ["   • Max LRU: Maximum threshold (system alerts when at/over this)", ""],
                ["   • Current LRU: Optional starting count (defaults to 0 if empty)", ""],
                ["   • Notes: Optional description or location details", ""],
                ["", ""],
                ["2. Guidelines", ""],
                ["   • Do NOT modify the header row (row 3 in Station Setup sheet)", ""],
                ["   • Station names must be unique", ""],
                ["   • Min and Max must be positive numbers", ""],
                ["   • Min must be less than Max", ""],
                ["   • Delete or clear the example rows before importing", ""],
                ["", ""],
                ["3. Import Process", ""],
                ["   • Save this file after filling in your data", ""],
                ["   • In the LRU Tracker app, click 'Import from Template'", ""],
                ["   • Select this file", ""],
                ["   • Review the import summary", ""],
                ["   • Confirm to add all stations", ""],
                ["", ""],
                ["4. Tips", ""],
                ["   • You can copy/paste data from existing spreadsheets", ""],
                ["   • Leave Current LRU empty if you'll update it later in the app", ""],
                ["   • Use the Notes column for any special instructions", ""],
                ["   • Keep this template for future use", ""],
                ["", ""],
                ["5. Example Values", ""],
                ["   Station Type          | Typical Min | Typical Max", ""],
                ["   Pack Stations         | 5-10        | 15-25", ""],
                ["   Dock Doors           | 10-15       | 25-40", ""],
                ["   Induct Stations      | 3-5         | 10-20", ""],
                ["   Palletize Zones      | 8-12        | 20-30", ""],
            ]
            
            for row_num, (col1, col2) in enumerate(instruction_text, 1):
                ws_instructions.cell(row_num, 1).value = col1
                ws_instructions.cell(row_num, 2).value = col2
                
                # Style the title
                if row_num == 1:
                    ws_instructions.cell(row_num, 1).font = Font(size=14, bold=True, color="2C3E50")
                
                # Style section headers
                if any(heading in col1 for heading in ["How to Use", "Guidelines", "Import Process", "Tips", "Example Values"]):
                    ws_instructions.cell(row_num, 1).font = Font(bold=True, size=11)
            
            ws_instructions.column_dimensions['A'].width = 50
            ws_instructions.column_dimensions['B'].width = 40
            
            wb.save(filename)
            messagebox.showinfo("Success", 
                              f"Template downloaded successfully!\n\n"
                              f"Location: {filename}\n\n"
                              f"Fill in your stations and use 'Import from Template' to load them.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create template:\n{str(e)}")
    
    def import_from_template(self):
        """Import stations from a filled template Excel file"""
        filename = filedialog.askopenfilename(
            title="Select Station Template to Import",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.load_workbook(filename)
            
            # Try to find the Station Setup sheet
            if "Station Setup" in wb.sheetnames:
                ws: Worksheet = wb["Station Setup"]  # type: ignore
            else:
                ws: Worksheet = wb.active  # type: ignore
            
            imported_stations = []
            errors = []
            skipped = []
            updated = []
            
            # Find header row (look for "Station Name")
            header_row = None
            for row in range(1, 10):
                cell_value = ws.cell(row, 1).value
                if cell_value and "Station Name" in str(cell_value):
                    header_row = row
                    break
            
            if not header_row:
                messagebox.showerror("Error", 
                                   "Could not find header row with 'Station Name'.\n"
                                   "Please use the downloaded template format.")
                return
            
            # Read data starting after header
            for row_num in range(header_row + 1, ws.max_row + 1):
                station_name = ws.cell(row_num, 1).value
                min_lru = ws.cell(row_num, 2).value
                max_lru = ws.cell(row_num, 3).value
                current_lru = ws.cell(row_num, 4).value
                notes = ws.cell(row_num, 5).value
                
                # Skip empty rows
                if not station_name or str(station_name).strip() == "":
                    continue
                
                station_name = str(station_name).strip()
                
                # Validate data
                try:
                    min_val = int(str(min_lru)) if (min_lru and str(min_lru).strip()) else 5
                    max_val = int(str(max_lru)) if (max_lru and str(max_lru).strip()) else 20
                    current_val = int(str(current_lru)) if (current_lru and str(current_lru).strip()) else 0
                    
                    if min_val < 0 or max_val < 0 or current_val < 0:
                        errors.append(f"Row {row_num}: Negative values not allowed for '{station_name}'")
                        continue
                    
                    if min_val > max_val:
                        errors.append(f"Row {row_num}: Min ({min_val}) > Max ({max_val}) for '{station_name}'")
                        continue
                    
                    # Check if station already exists
                    if station_name in self.stations:
                        # Ask user what to do
                        action = messagebox.askyesnocancel(
                            "Station Exists",
                            f"Station '{station_name}' already exists.\n\n"
                            f"Current: Min={self.stations[station_name]['min']}, "
                            f"Max={self.stations[station_name]['max']}\n"
                            f"Template: Min={min_val}, Max={max_val}\n\n"
                            f"Yes = Update existing\n"
                            f"No = Skip this station\n"
                            f"Cancel = Stop import"
                        )
                        
                        if action is None:  # Cancel
                            messagebox.showinfo("Import Cancelled", "Import process was cancelled by user.")
                            return
                        elif action:  # Yes - Update
                            self.stations[station_name]['min'] = min_val
                            self.stations[station_name]['max'] = max_val
                            if current_val > 0:
                                self.stations[station_name]['current'] = current_val
                            updated.append(station_name)
                        else:  # No - Skip
                            skipped.append(station_name)
                        continue
                    
                    # Add new station
                    self.stations[station_name] = {
                        'current': current_val,
                        'min': min_val,
                        'max': max_val,
                        'history': []
                    }
                    
                    # Add initial history entry if current value provided
                    if current_val > 0:
                        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        self.stations[station_name]['history'].append({
                            'timestamp': timestamp,
                            'count': current_val
                        })
                        self.history.append({
                            'station': station_name,
                            'timestamp': timestamp,
                            'count': current_val,
                            'min': min_val,
                            'max': max_val
                        })
                    
                    imported_stations.append(station_name)
                    
                except (ValueError, TypeError) as e:
                    errors.append(f"Row {row_num}: Invalid data for '{station_name}' - {str(e)}")
                    continue
            
            # Save and refresh
            if imported_stations or updated:
                self.save_data()
                self.refresh_display()
            
            # Show summary
            summary = f"Import Complete!\n\n"
            summary += f"✅ Imported: {len(imported_stations)} stations\n"
            if updated:
                summary += f"🔄 Updated: {len(updated)} stations\n"
            if skipped:
                summary += f"⏭️ Skipped: {len(skipped)} stations\n"
            if errors:
                summary += f"❌ Errors: {len(errors)}\n"
            
            if imported_stations:
                summary += f"\n📋 Imported Stations:\n" + "\n".join(f"  • {s}" for s in imported_stations[:10])
                if len(imported_stations) > 10:
                    summary += f"\n  ... and {len(imported_stations) - 10} more"
            
            if errors:
                summary += f"\n\n⚠️ Errors:\n" + "\n".join(errors[:5])
                if len(errors) > 5:
                    summary += f"\n... and {len(errors) - 5} more errors"
            
            messagebox.showinfo("Import Summary", summary)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to import template:\n{str(e)}")
    
    def import_fc_schedule(self):
        """Import stations from FC Standard Work Spreadsheet CSV format"""
        filename = filedialog.askopenfilename(
            title="Select FC Standard Work Spreadsheet CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        if not filename:
            return
        
        try:
            import csv
            
            imported_stations = []
            errors = []
            skipped = []
            updated = []
            
            # Read CSV file
            with open(filename, 'r', encoding='utf-8-sig') as f:
                csv_reader = csv.reader(f)
                rows = list(csv_reader)
            
            # Parse the FC schedule format
            # Looking for rows with: LRU, Test Description, Rack Location, time slots...
            for row_num, row in enumerate(rows, 1):
                # Skip empty rows or header rows
                if len(row) < 3:
                    continue
                
                lru_name = str(row[0]).strip() if row[0] else ""
                test_desc = str(row[1]).strip() if row[1] else ""
                rack_location = str(row[2]).strip() if row[2] else ""
                
                # Skip if LRU name is empty or looks like a header
                if not lru_name or "LRU" in lru_name or "Shift" in lru_name:
                    continue
                
                # Create station name from LRU + Rack Location
                if rack_location:
                    station_name = f"{lru_name} - {rack_location}"
                else:
                    station_name = lru_name
                
                # Extract batch size from test description (B=X format)
                import re
                batch_match = re.search(r'B\s*=\s*(\d+)', test_desc, re.IGNORECASE)
                
                if batch_match:
                    batch_size = int(batch_match.group(1))
                    # Use batch size to set reasonable min/max
                    min_val = max(1, batch_size // 2)  # Min = half batch
                    max_val = batch_size * 2  # Max = double batch
                else:
                    # Default values if no batch size found
                    min_val = 5
                    max_val = 20
                
                # Check if station already exists
                if station_name in self.stations:
                    action = messagebox.askyesnocancel(
                        "Station Exists",
                        f"Station '{station_name}' already exists.\n\n"
                        f"Current: Min={self.stations[station_name]['min']}, "
                        f"Max={self.stations[station_name]['max']}\n"
                        f"CSV suggests: Min={min_val}, Max={max_val}\n\n"
                        f"Yes = Update existing\n"
                        f"No = Skip this station\n"
                        f"Cancel = Stop import"
                    )
                    
                    if action is None:  # Cancel
                        messagebox.showinfo("Import Cancelled", "Import process was cancelled by user.")
                        return
                    elif action:  # Yes - Update
                        self.stations[station_name]['min'] = min_val
                        self.stations[station_name]['max'] = max_val
                        updated.append(station_name)
                    else:  # No - Skip
                        skipped.append(station_name)
                    continue
                
                # Add new station
                self.stations[station_name] = {
                    'current': 0,
                    'min': min_val,
                    'max': max_val,
                    'history': [],
                    'test_description': test_desc,  # Store for reference
                    'rack_location': rack_location
                }
                
                imported_stations.append(station_name)
            
            # Save and refresh
            if imported_stations or updated:
                self.save_data()
                self.refresh_display()
            
            # Show detailed summary
            summary = f"FC Schedule Import Complete!\n\n"
            summary += f"✅ Imported: {len(imported_stations)} stations\n"
            if updated:
                summary += f"🔄 Updated: {len(updated)} stations\n"
            if skipped:
                summary += f"⏭️ Skipped: {len(skipped)} stations\n"
            
            if imported_stations:
                summary += f"\n📋 Imported Stations:\n"
                for station in imported_stations[:15]:
                    station_data = self.stations[station]
                    summary += f"  • {station}\n    Min: {station_data['min']}, Max: {station_data['max']}\n"
                if len(imported_stations) > 15:
                    summary += f"  ... and {len(imported_stations) - 15} more\n"
            
            summary += f"\n💡 Tip: Min/Max values were auto-calculated from batch sizes (B=X).\n"
            summary += f"You can edit any station to adjust these values."
            
            messagebox.showinfo("Import Summary", summary)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to import FC schedule:\n{str(e)}\n\nMake sure the file is a CSV format.")
    
    def check_for_updates(self):
        """Check for application updates with security validation"""
        try:
            import urllib.request
            import urllib.error
            
            # Security check: Validate URL is HTTPS
            if not UPDATE_CHECK_URL.startswith('https://'):
                messagebox.showwarning("Security Warning", 
                    "Update URL must use HTTPS for security.\n"
                    "Update checking is disabled for safety.")
                return
            
            # Show checking message
            checking_window = tk.Toplevel(self.root)
            checking_window.title("Checking for Updates")
            checking_window.geometry("300x100")
            checking_window.transient(self.root)
            checking_window.grab_set()
            
            tk.Label(checking_window, text="🔍 Checking for updates...", 
                    font=('Arial', 12)).pack(pady=30)
            checking_window.update()
            
            def check_updates_thread():
                try:
                    # Download version info with timeout
                    with urllib.request.urlopen(UPDATE_CHECK_URL, timeout=5) as response:
                        content = response.read().decode('utf-8')
                        
                        # Validate JSON before parsing
                        try:
                            update_info = json.loads(content)
                        except json.JSONDecodeError:
                            raise ValueError("Invalid update data received")
                        
                        # Validate required fields
                        if not isinstance(update_info, dict):
                            raise ValueError("Invalid update data format")
                        
                        required_fields = ['version', 'download_url']
                        missing_fields = [f for f in required_fields if f not in update_info]
                        if missing_fields:
                            raise ValueError(f"Missing required fields: {', '.join(missing_fields)}")
                    
                    latest_version = update_info.get('version', '0.0.0')
                    
                    # Validate version format
                    if not self._validate_version_format(latest_version):
                        raise ValueError("Invalid version format")
                    
                    # Close checking window
                    checking_window.destroy()
                    
                    # Compare versions
                    if self._is_newer_version(latest_version, APP_VERSION):
                        # Show update available dialog
                        self._show_update_dialog(update_info)
                    else:
                        messagebox.showinfo("No Updates", 
                                          f"You're running the latest version ({APP_VERSION})!")
                
                except (urllib.error.URLError, urllib.error.HTTPError) as e:
                    checking_window.destroy()
                    messagebox.showwarning("Network Error", 
                                         "Could not connect to update server.\n\n"
                                         "Please check your internet connection.")
                except ValueError as e:
                    checking_window.destroy()
                    messagebox.showerror("Invalid Update Data", 
                                       f"The update information is invalid or corrupted.\n\n"
                                       f"Please try again later or contact support.")
                except Exception as e:
                    checking_window.destroy()
                    messagebox.showerror("Update Check Failed", 
                                       "An unexpected error occurred while checking for updates.")
            
            # Run in thread to avoid freezing UI
            thread = threading.Thread(target=check_updates_thread, daemon=True)
            thread.start()
            
        except Exception as e:
            messagebox.showerror("Error", "Update check failed to start.")
    
    def _validate_version_format(self, version: str) -> bool:
        """Validate version string format (e.g., 1.0.0)"""
        import re
        return bool(re.match(r'^\d+\.\d+\.\d+$', version))
    
    def _is_newer_version(self, latest: str, current: str) -> bool:
        """Compare version strings securely"""
        try:
            # Validate both version strings
            if not (self._validate_version_format(latest) and 
                    self._validate_version_format(current)):
                return False
            
            latest_parts = [int(x) for x in latest.split('.')]
            current_parts = [int(x) for x in current.split('.')]
            return latest_parts > current_parts
        except (ValueError, AttributeError):
            return False
    
    def _show_update_dialog(self, update_info: dict):
        """Show dialog with update information"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Update Available")
        dialog.geometry("500x400")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Title
        title_frame = tk.Frame(dialog, bg='#27ae60')
        title_frame.pack(fill='x')
        
        tk.Label(title_frame, text="🎉 Update Available!", 
                font=('Arial', 16, 'bold'), bg='#27ae60', fg='white',
                pady=15).pack()
        
        # Content
        content_frame = tk.Frame(dialog, bg='white')
        content_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Version info
        tk.Label(content_frame, text=f"Current Version: {APP_VERSION}", 
                font=('Arial', 11), bg='white').pack(anchor='w')
        tk.Label(content_frame, text=f"Latest Version: {update_info.get('version', 'Unknown')}", 
                font=('Arial', 11, 'bold'), bg='white', fg='#27ae60').pack(anchor='w', pady=(0, 15))
        
        # Release notes
        if update_info.get('release_notes'):
            tk.Label(content_frame, text="What's New:", 
                    font=('Arial', 11, 'bold'), bg='white').pack(anchor='w', pady=(5, 5))
            
            notes_frame = tk.Frame(content_frame, bg='#ecf0f1', relief='sunken', bd=1)
            notes_frame.pack(fill='both', expand=True, pady=(0, 15))
            
            notes_text = tk.Text(notes_frame, wrap='word', height=8, 
                                font=('Arial', 10), bg='#ecf0f1', relief='flat')
            notes_text.pack(fill='both', expand=True, padx=10, pady=10)
            notes_text.insert('1.0', update_info.get('release_notes', ''))
            notes_text.config(state='disabled')
        
        # Size info
        if update_info.get('size_mb'):
            tk.Label(content_frame, text=f"Download Size: ~{update_info.get('size_mb')} MB", 
                    font=('Arial', 9), bg='white', fg='#7f8c8d').pack(anchor='w')
        
        # Buttons
        button_frame = tk.Frame(dialog, bg='white')
        button_frame.pack(fill='x', padx=20, pady=(0, 20))
        
        def open_download():
            import webbrowser
            webbrowser.open(update_info.get('download_url', ''))
            dialog.destroy()
        
        tk.Button(button_frame, text="📥 Download Update", 
                 command=open_download,
                 bg='#27ae60', fg='white', font=('Arial', 11, 'bold'),
                 padx=20, pady=10).pack(side='left', padx=(0, 10))
        
        tk.Button(button_frame, text="Later", 
                 command=dialog.destroy,
                 bg='#95a5a6', fg='white', font=('Arial', 11),
                 padx=20, pady=10).pack(side='left')

def main():
    root = tk.Tk()
    app = LRUTrackerApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
